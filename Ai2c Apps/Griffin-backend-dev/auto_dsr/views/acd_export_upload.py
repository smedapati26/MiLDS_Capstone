from http import HTTPStatus

import pandas as pd
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import load_workbook

from auto_dsr.models import ACDExport, Unit, User
from auto_dsr.utils.acd_export.update_1352 import update_1352s
from auto_dsr.utils.acd_export.update_item import update_item
from utils.http import (
    HTTP_ERROR_MESSAGE_DAILY_STATUS_MALFORMED,
    HTTP_ERROR_MESSAGE_INSPECTION_MALFORMED,
    HTTP_ERROR_MESSAGE_NO_FILE,
    HTTP_ERROR_MESSAGE_PHASE_MALFORMED,
    HTTP_ERROR_MESSAGE_READINESS_MALFORMED,
    HTTP_ERROR_MESSAGE_READINESS_STATUS_MALFORMED,
    HTTP_ERROR_MESSAGE_UNIT_DOES_NOT_EXIST,
    HTTP_ERROR_MESSAGE_USER_DOES_NOT_EXIST,
    get_user_id,
)
from utils.time import get_reporting_period


@require_POST
def acd_export_upload(request: HttpRequest) -> HttpResponse:
    """
    Handles updating all aircraft contained within an ACD export upload

    @param request: (django.http.HttpRequest) the request object
    @param uic: (str) the uic of the unit selected when the file was uploaded
    """
    error_string = ""
    shiny_upload = request.headers["User-Agent"].startswith("libcurl")

    try:  # to get the user uploading the export
        if shiny_upload:
            user_id = request.headers.get("X-On-Behalf-Of", None)
            user = User.objects.get(user_id=user_id)
        else:  # Uploaded directly from small upload link
            user_id = get_user_id(request.headers)
            user = User.objects.get(user_id=user_id)
    except User.DoesNotExist:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_USER_DOES_NOT_EXIST}, status=HTTPStatus.NOT_FOUND)
    except KeyError:
        return JsonResponse(
            {"error": "Improperly formatted request, no user id passed in headers."}, status=HTTPStatus.BAD_REQUEST
        )

    try:  # to get the unit being updated by the export
        unit = Unit.objects.get(uic=request.POST.get("unit"))
    except Unit.DoesNotExist:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_UNIT_DOES_NOT_EXIST}, status=HTTPStatus.NOT_FOUND)

    dsr_export_time = timezone.now().replace(microsecond=0)
    export = ACDExport.objects.create(unit=unit, user=user, uploaded_at=dsr_export_time, upload_type="xlsx")

    try:  # to get the export from the request
        export_file = request.FILES["acd_export"]
        export.document = export_file
        export.save()
    except KeyError:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_NO_FILE}, status=HTTPStatus.BAD_REQUEST)
    except Exception as e:
        print("Document saving failed, proceding with updates...")

    sync = request.POST.get("sync", "FALSE")
    overwrite_pauses = False if sync == "FALSE" else True  # Shiny Material Switch sends text instead of logical bool

    try:  # to load the Workbook
        wb = load_workbook(export_file, data_only=True)
    except:
        return JsonResponse(
            {"error": "Failed To Read Workbook. Please remove macros or other special formatting from the file."},
            status=HTTPStatus.BAD_REQUEST,
        )

    try:
        daily_status_page = wb["Daily Status"]
        aircraft_count = daily_status_page["B2"].value
        if aircraft_count == 0:
            return JsonResponse(
                {
                    "error": """Improperly created export file. Please revisit the instructions and
                    ensure you set the DailyStatus-DailyDate filter in your ACD when generating the export."""
                },
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        header_offset = 4
        dsr_data = list(
            daily_status_page.iter_rows(
                min_row=header_offset,
                max_col=18,
                max_row=aircraft_count + header_offset,
                values_only=True,
            )
        )
        dsr_df = pd.DataFrame.from_records(dsr_data[1:], columns=dsr_data[0], index="DailyStatus-End Item")
    except:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_DAILY_STATUS_MALFORMED}, status=HTTPStatus.BAD_REQUEST)

    try:
        readiness_page = wb["Readiness"]
        readiness_records_count = readiness_page["B2"].value
        if readiness_records_count == 0:
            error_string += "Readiness Data Not Added. "
            '''return JsonResponse(
                {
                    "error": """Improperly created export file. Please revisit the instructions and
                    ensure you set the Readiness-PeriodEnding and ReadinessStatus-PeriodEnding filters in
                    your ACD when generating the export."""
                },
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )'''
        readiness_data = list(
            readiness_page.iter_rows(
                min_row=header_offset,
                max_col=20,
                max_row=readiness_records_count + header_offset,
                values_only=True,
            )
        )
        readiness_df = pd.DataFrame.from_records(
            readiness_data[1:], columns=readiness_data[0], index="Readiness-End Item"
        )
        start_date, _ = get_reporting_period()
        if readiness_df.empty:
            readiness_flight_hours = pd.DataFrame(
                columns=["Readiness-End Item", "Readiness-Flying Hours", "Readiness-Total Airframe Hours"]
            )
        else:
            readiness_flight_hours = (
                readiness_df[readiness_df["Readiness-Day"].dt.date >= start_date]
                .groupby("Readiness-End Item")
                .agg({"Readiness-Flying Hours": "sum", "Readiness-Total Airframe Hours": "max"})
            )
        all_data = dsr_df.join(readiness_flight_hours, how="left").fillna(
            {"Readiness-Flying Hours": 0.0, "Readiness-Total Airframe Hours": 0.0}
        )
    except:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_READINESS_MALFORMED}, status=HTTPStatus.BAD_REQUEST)

    try:
        inspections_page = wb["Inspection Schedule"]
        inspections_count = inspections_page["B2"].value
        inspections_data = list(
            inspections_page.iter_rows(
                min_row=header_offset,
                max_row=inspections_count + header_offset,
                max_col=21,
                values_only=True,
            )
        )
        inspections_df = pd.DataFrame.from_records(inspections_data[1:], columns=inspections_data[0])
    except:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_INSPECTION_MALFORMED}, status=HTTPStatus.BAD_REQUEST)

    try:
        phase_data_page = wb["Phase Flow Chart"]
        phase_data_count = phase_data_page["B2"].value
        phase_data = list(
            phase_data_page.iter_rows(
                min_row=header_offset, max_row=phase_data_count + header_offset, max_col=11, values_only=True
            )
        )
        phase_df = pd.DataFrame.from_records(phase_data[1:], columns=phase_data[0], index="PhaseFlowChart-End Item")
    except:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_PHASE_MALFORMED}, status=HTTPStatus.BAD_REQUEST)

    try:
        all_data.apply(
            lambda row: update_item(
                row,
                dsr_export_time,
                inspections_df[inspections_df["InspectionSchedule-End Item"] == row.name],
                phase_df[phase_df.index == row.name],
                overwrite_pauses,
            ),
            axis=1,
        )
    except:
        return JsonResponse({"error": "Failed to update items"}, status=HTTPStatus.BAD_REQUEST)

    try:
        readiness_status_page = wb["Readiness Status"]
        readiness_status_records_count = readiness_status_page["B2"].value
        if readiness_status_records_count == 0:
            error_string += """1352 Data not updated. If using ACN workaround, this behavior is expected. 
                            It is encouraged to manually check flight hours. If not using ACN workaround
                            please notify the Griffin team via the support channel"""
            return JsonResponse(
                {"error": error_string},
                status=HTTPStatus.PARTIAL_CONTENT,
            )
        readiness_status_data = list(
            readiness_status_page.iter_rows(
                min_row=header_offset,
                max_col=11,
                max_row=readiness_status_records_count + header_offset,
                values_only=True,
            )
        )
        readiness_status_df = pd.DataFrame.from_records(
            readiness_status_data[1:], columns=readiness_status_data[0], index="ReadinessStatus-End Item"
        )
        readiness_period = readiness_status_df["ReadinessStatus-Period Ending"].iloc[0]
    except:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_READINESS_STATUS_MALFORMED}, status=HTTPStatus.BAD_REQUEST)

    try:
        update_1352s(readiness_status_df, readiness_period)
    except:
        return JsonResponse(
            {
                "error": """DSR data updated but 1352 data update failed due to improperly created export file.
                Please revisit the instructions and ensure you set the ReadinessStatus-PeriodEnding filter in
                your ACD when generating the export."""
            },
            status=HTTPStatus.PARTIAL_CONTENT,
        )

    return HttpResponse("Read File")
