from http import HTTPStatus

import pandas as pd
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import load_workbook

from auto_dsr.models import DA2407Export, Unit, User
from auto_dsr.utils.acd_export.update_da_2407 import update_da_2407s
from utils.http import (
    HTTP_ERROR_MESSAGE_DA_2407_MALFORMED,
    HTTP_ERROR_MESSAGE_NO_FILE,
    HTTP_ERROR_MESSAGE_UNIT_DOES_NOT_EXIST,
    HTTP_ERROR_MESSAGE_USER_DOES_NOT_EXIST,
    get_user_id,
)


@require_POST
def da_2407_export_upload(request: HttpRequest) -> HttpResponse:
    """
    Handles adding all da_2407s from an export file to the db.

    @param request: (django.http.HttpRequest) the request object
    @param uic: (str) the uic of the unit selected when the file was uploaded
    """
    error_string = ""
    shiny_upload = request.headers["User-Agent"].startswith("libcurl")

    try:  # to get the user uploading the export
        if shiny_upload:
            user_id = request.headers.get("X-On-Behalf-Of", None)
            user = User.objects.get(user_id=user_id)
        else:  # Uploaded directly from small upload link
            user_id = get_user_id(request.headers)
            user = User.objects.get(user_id=user_id)
    except User.DoesNotExist:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_USER_DOES_NOT_EXIST}, status=HTTPStatus.NOT_FOUND)
    except KeyError:
        return JsonResponse(
            {"error": "Improperly formatted request, no user id passed in headers."}, status=HTTPStatus.BAD_REQUEST
        )

    try:  # to get the unit being updated by the export
        unit = Unit.objects.get(uic=request.POST.get("unit"))
    except Unit.DoesNotExist:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_UNIT_DOES_NOT_EXIST}, status=HTTPStatus.NOT_FOUND)

    da_2407_export_time = timezone.now().replace(microsecond=0)
    export = DA2407Export.objects.create(unit=unit, user=user, uploaded_at=da_2407_export_time, upload_type="xlsx")

    try:  # to get the export from the request
        export_file = request.FILES["da_2407_export"]
        export.document = export_file
        export.save()
    except KeyError:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_NO_FILE}, status=HTTPStatus.BAD_REQUEST)
    except Exception as e:
        print("Document saving failed, proceding with updates...")

    try:  # to load the Workbook
        wb = load_workbook(export_file, data_only=True)
    except:
        return JsonResponse(
            {"error": "Failed To Read Workbook. Please remove macros or other special formatting from the file."},
            status=HTTPStatus.BAD_REQUEST,
        )

    try:
        work_order_page = wb["2407 Detail"]
        work_order_count = work_order_page["B2"].value
        if work_order_count == 0:
            return JsonResponse(
                {
                    "error": """Improperly created export file or no active work orders. Please revisit the instructions and
                    ensure you set the correct filters in ACD when generating the export."""
                },
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        header_offset = 4
        work_order_data = list(
            work_order_page.iter_rows(
                min_row=header_offset,
                max_col=78,
                max_row=work_order_count + header_offset,
                values_only=True,
            )
        )
        work_order_df = pd.DataFrame.from_records(work_order_data[1:], columns=work_order_data[0], index="UIC WON")
    except:
        return JsonResponse({"error": HTTP_ERROR_MESSAGE_DA_2407_MALFORMED}, status=HTTPStatus.BAD_REQUEST)

    try:
        update_da_2407s(work_order_df)
    except:
        return JsonResponse(
            {
                "error": """Work order updates failed due to improperly created export file.
                Please revisit the instructions and ensure you set the correct filter in
                your ACD when generating the export."""
            },
            status=HTTPStatus.PARTIAL_CONTENT,
        )

    return HttpResponse("Read File")
